# PROGRAMA EM PYTHON PARA FAZER COMPARAÇÃO DE TABELAS EM ARQUIVOS ACCESS .accdb
# Autor: Rafael Henrique da Rosa
# Estagiário Divisão de Engenharia de Manutenção
# Eletrônica (SMIN.DT) - Itaipu Binacional
# Abril de 2022

# TODO:*
# FEITO Filtragem de texto (pesquisa)
# FEITO Verificar o tamanho do arquivo para mudar o tipo de importação
# FEITO tirar o filtro
# FEITO Retirar o ponto nos inteiros (existe uma função comentada no codigo
#       mas é pessima para tabelas grandes)
# FEITO Criar um executavel sem a pasta
# FEITO Verificar se a tabela existe nos dois bancos
# FEITO Verificar se o arquivo novo e antigo não é o mesmo
# FEITO Arrumar a exportação do relatório
# FEITO Fazer uma logica de resize e move

from calendar import c
import subprocess
import pandas as pd
import os
import tkinter as tk
import openpyxl
from tkinter import ttk, messagebox, Frame, Label, Entry, Toplevel
from tkinter import filedialog as fd

from pandastable import Table
from pandastable import config
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import multiprocessing
import regex as re
import sys

global colore


global path1
global path2

path1 = ''
path2 = ''
selected_table = ""
campos = ['Nenhum', 'Nenhum', 'Nenhum']
colunas = []
table1 = pd.DataFrame()


def resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def pinta_discrep():

    # Função responsavel por mudar a cor das celulas
    # que possuem os mesmos valores de comparação
    # e outros valores diferentes
    # ESSA FUNÇÃO DOBRA O TEMPO DE COMPILAÇÃO
    vermelho = '#ff0000'
    # laranja = '#ff8400'
    vermelho_claro = '#fa9898'
    verde_claro = '#98faa7'
    # passa pelo dataframe de discrepantes com passo de 2
    for i in range(0, table_discrep.shape[0]-1, 2):
        # obtem uma lista de booleanos indicando valores diferentes
        uq = table_discrep.iloc[i] != table_discrep.iloc[i+1]
        uq = uq.to_list()
        # obtem uma lista de booleanos indicando campos com "NaN"
        # Necessario pois por algum motivo a comparação NaN = NaN retorna False
        uq2 = table_discrep.iloc[i].isnull()
        uq2 = uq2.to_list()
        # percorre as listas e pinta de vermelho os valores diferentes
        for j in range(1, len(uq)):
            if uq[j] is True and uq2[j] is False:
                pt_resul_discrep.setRowColors(rows=i, clr=vermelho, cols=[j])
                pt_resul_discrep.setRowColors(rows=i+1,
                                              clr=vermelho, cols=[j])

    # Pinta os elementos discrepantes da tabela antiga
    indexes = table_discrep.index.tolist()
    for i in range(table_discrep.shape[0]):
        if i % 2 == 0:
            pt1.setRowColors(rows=indexes[i]-1, clr=vermelho, cols=[0, 1])
        else:
            pt2.setRowColors(rows=indexes[i]-1, clr=vermelho, cols=[0, 1])

    # Pinta os elementos novos na tabela nova
    indexes = table_novas.index.tolist()
    for i in range(table_novas.shape[0]):
        pt2.setRowColors(rows=indexes[i]-1, clr=vermelho_claro, cols=[0, 1])

    # Pinta os elementos excluidos na tabela antiga
    indexes = table_excluidas.index.tolist()
    for i in range(table_excluidas.shape[0]):
        pt1.setRowColors(rows=indexes[i]-1, clr=verde_claro, cols=[0, 1])


def clear_table():

    pt_resul_discrep.model.df = df
    pt_resul_discrep.redraw()
    pt_resul_discrep.autoResizeColumns()
    pt_resul_discrep.show()
    pt_resul_discrep.showIndex()
    pt_resul_discrep.redraw()
    lbl_discrep.configure(text="LINHAS DISCREPANTES: ")

    pt_resul_novas.model.df = df
    pt_resul_novas.redraw()
    pt_resul_novas.autoResizeColumns()
    pt_resul_novas.show()
    pt_resul_novas.showIndex()
    pt_resul_novas.redraw()
    lbl_novas.configure(text="LINHAS ADICIONADAS (presentes somente no "
                             "arquivo novo): ")

    pt_resul_excluidas.model.df = df
    pt_resul_excluidas.redraw()
    pt_resul_excluidas.autoResizeColumns()
    pt_resul_excluidas.show()
    pt_resul_excluidas.redraw()
    pt_resul_excluidas.showIndex()
    lbl_excluidas.configure(text="LINHAS EXCLUIDAS (presentes somente no"
                            " arquivo antigo): ")


def update_table(filtrado):
    # Função responsavel por atualizar as tabelas na interface após comparação

    # Atualiza a tabela 1
    if filtrado:
        pt1.model.df = table1_filtrada
    else:
        pt1.model.df = table1
    pt1.autoResizeColumns()
    pt1.redraw()
    pt1.show()

    # Atualiza a tabela 2
    if filtrado:
        pt2.model.df = table2_filtrada
    else:
        pt2.model.df = table2
    pt2.autoResizeColumns()
    pt2.redraw()
    pt2.show()

    # Atualiza a tabela de discrepancias
    if filtrado:
        pt_resul_discrep.model.df = table_discrep_filtrada
    else:
        pt_resul_discrep.model.df = table_discrep

    lbl_discrep.configure(text="LINHAS DISCREPANTES: "
                          + str(table_discrep1.shape[0]))

    pt_resul_discrep.autoResizeColumns()
    pt_resul_discrep.show()
    pt_resul_discrep.showIndex()
    pt_resul_discrep.redraw()
    pt_resul_discrep.autoResizeColumns()
    pt_resul_discrep.show()
    # Função que pinta os valores discrepantes

    global colore

    if colore.get() is True:
        pinta_discrep()
    # Deixa uma linha sem dados como a selecionada
    # A linha selecionada tem uma cor diferente
    # e não mostra os valores discrepantes pintados
# =============================================================================
# =============================================================================
# #     #pt_resul_discrep.movetoSelection(row=table_discrep.shape[0]+1, col=0)
# =============================================================================
# =============================================================================
    pt_resul_discrep.redraw()

    # Atualiza a tabela das linhas novas
    lbl_novas.configure(text="LINHAS ADICIONADAS (presentes somente no "
                        "arquivo novo): " + str(table_novas.shape[0])),
    if filtrado:
        pt_resul_novas.model.df = table_novas_filtrada
    else:
        pt_resul_novas.model.df = table_novas
    pt_resul_novas.autoResizeColumns()
    pt_resul_novas.show()
    pt_resul_novas.showIndex()
    pt_resul_novas.redraw()
    pt_resul_novas.autoResizeColumns()
    pt_resul_novas.show() 
    # Atualiza a tabela das linhas excluidas
    lbl_excluidas.configure(text="LINHAS EXCLUIDAS (presentes somente no"
                            " arquivo antigo): "
                            + str(table_excluidas.shape[0]))
    if filtrado:
        pt_resul_excluidas.model.df = table_excluidas_filtrada
    else:
        pt_resul_excluidas.model.df = table_excluidas

    pt_resul_excluidas.autoResizeColumns()
    pt_resul_excluidas.show()
    pt_resul_excluidas.showIndex()
    pt_resul_excluidas.redraw()
    pt_resul_excluidas.autoResizeColumns()
    pt_resul_excluidas.show()


def process_importa_antigo(path, file1, selected_table):
    # Cria uma linha de comando e executa no cmd para importar o arquivo antigo
    # Funciona somente se mdb-export.exe existe na pasta mdbtools
    export_command = path
    export_command += ' ' + file1
    export_command += ' '
    export_command += selected_table + '  > temp1.csv'
    export1 = []
    #export1.append('cmd.exe')
    #export1.append('/c')
    export1.append(path)
    export1.append(file1)
    export1.append(selected_table)
    export1.append(">")
    export1.append("temp1.csv")

    #print(export1)
    # executa a linha de comando no cmd
    subprocess.run(export1, shell=True)



def process_importa_novo(path, file2, selected_table):
    # Cria uma linha de comando e executa no cmd para importar o arquivo novo
    # Funciona somente se mdb-export.exe existe na pasta mdbtools
    export_command = path
    export_command += ' ' + file2
    export_command += ' '
    export_command += selected_table + '  > temp2.csv'
    export2 = []
    export2.append(path)
    export2.append(file2)
    export2.append(selected_table)
    export2.append(">")
    export2.append("temp2.csv")

    #print(export2)
    # subprocess.run(['cmd.exe', '/c', export_command])
    subprocess.run(export2, shell=True)


def load_tables():
    global table_novas
    global table_excluidas
    global table_discrep
    global table_discrep1
    global table1
    global table2
    global selected_table

    file1 = path1
    file2 = path2
    file_temp1 = path1
    file_temp2 = path2
    export1= []
    export2 = []
    # Caso a comparação seja em arquivos access:
    if file1.endswith('.accdb'):
        
        # Corrige o nome do arquivo para funcionar o mdb-tools
        # O nome do arquivo não pode ter espaço
        # file =   file1.split("/")
        # tmp = file[len(file)-1]

        # if " " in file[len(file)-1]:
        #     file_temp1 =""
        #     file[len(file)-1] = file[len(file)-1].replace(" ", "_")
        #     print(file[len(file)-1])
        #     for i in range(len(file)):
        #         file_temp1+= file[i]
        #         if i != len(file)-1:
        #             file_temp1+= "/"
        #     os.rename(file1, file_temp1)
            


        # file =   file2.split("/")

        # if " " in file2:
        #     file_temp2 =""
        #     file[len(file)-1] = file[len(file)-1].replace(" ", "_")
        #     for i in range(len(file)):
        #         file_temp2+= file[i]
        #         if i != len(file)-1:
        #             file_temp2+= "/"
        #     os.rename(file2, file_temp2)

        # Pega o numero de linhas em cada tabela
        path = resource_path('mdbtools\\mdb-count.exe')
        # test = path.split("\\")
        # path = ""
        # for i in range(len(test)):
        #     if " " in test[i]:
        #         test[i] = '"' + test[i] + '"'
        #     print(test[i])
        #     path+= test[i]
        #     if i != len(test)-1:
        #         path+= "/"
        # print("T=" +file_temp1)
        export1.append(path)
        export2.append(path)

        export_command = path
        export_command += ' "' + file_temp1
        export1.append(file_temp1)
        export2.append(file_temp2)

        export_command += '" '
        export_command += selected_table
        export1.append(selected_table)
        export2.append(selected_table)

        #print(export1)
        rows1 = subprocess.check_output(
                    export1).decode()
        
        #print("ROWS1 = " + rows1)

        export_command = path
        export_command += ' "' + file_temp2
        export_command += '" '
        export_command += selected_table
        rows2 = subprocess.check_output(
                    export2).decode()
        #print("ROWS2 = " + rows1)
        # Seleciona a pasta mdbtools que deve estar na mesma pasta do programa
        path = resource_path('mdbtools\\mdb-export.exe')
        # test = path.split("\\")
        # path = ""
        # for i in range(len(test)):
        #     if " " in test[i]:
        #         test[i] = '"' + test[i] + '"'
        #     print(test[i])
        #     path+= test[i]
        #     if i != len(test)-1:
        #         path+= "/"
        # print(path)
        if __name__ == '__main__':
            # Se as tabelas tiverem mais de 1000 linhas
            # importa por dois processos
            if int(rows1) > 1000 or int(rows2) > 1000:
                # Cria dois processos para importar os arquivos
                p1 = multiprocessing.Process(
                    target=process_importa_antigo,
                    args=(path, file_temp1, selected_table))
                p1.start()
                p2 = multiprocessing.Process(
                    target=process_importa_novo,
                    args=(path, file_temp2, selected_table))
                p2.start()
                # Espera a importação para continuar na main
                p1.join()
                p2.join()
            # Caso as tabelas sejam menores importa diretamente
            else:
                    export1 = []
                    export1.append(path)
                    export1.append(file_temp1)
                    export1.append(selected_table)
                    export1.append(">")
                    export1.append("temp1.csv")
                    # executa a linha de comando no cmd
                    subprocess.run(export1, shell=True)

                    export2 = []
                    export2.append(path)
                    export2.append(file_temp1)
                    export2.append(selected_table)
                    export2.append(">")
                    export2.append("temp2.csv")
                    # executa a linha de comando no cmd
                    subprocess.run(export2, shell=True)
        # importa o arquivo csv em um dataframe do pandas e exclui o arquivo
        # o encoding é necessário pois na tabela existe um caracter "°"
        try:
            table1 = pd.read_csv('temp1.csv', sep=',', encoding='iso-8859-1')
            os.remove("temp1.csv")
        except Exception:
            messagebox.showinfo(
                "ERRO", "nao encontrou aquivo csv1")

        # Exclui linhas vazias
        df2 = table1[table1.isna().all(axis=1)]
        for i in range(df2.shape[0]):
            table1 = table1.drop(df2.index[i])

        # importa o arquivo csv em um dataframe
        # do pandas e exclui o arquivo
        # o encoding é necessário pois na tabela existe um caracter "°"
        try:
            table2 = pd.read_csv('temp2.csv', sep=',', encoding='iso-8859-1')
            os.remove("temp2.csv")
        except Exception:
            messagebox.showinfo(
                "ERRO", "nao encontrou aquivo csv2")
        # print("tempo para importar " + str((time.time() - start)))
        # Exclui Linhas vazias
        df2 = table2[table2.isna().all(axis=1)]
        for i in range(df2.shape[0]):
            table2 = table2.drop(df2.index[i])

    # Caso a comparação seja em arquivos excel
    elif file1.endswith('.xlsx'):

        table1 = pd.read_excel(
            open(file_temp1, 'rb'), sheet_name=selected_table)
        table2 = pd.read_excel(
            open(file_temp2, 'rb'), sheet_name=selected_table)

    # ajusta o tipo de variáveis dos dataframes
    # APARENTEMENTE AUMENTA ABSURDAMENTE O TEMPO DE COMPARAÇÃO
    # table1 = table1.convert_dtypes()
    # table2 = table2.convert_dtypes()

    # Renomeia os arquivos para o nome original
    if(file_temp1 != file1):
        os.rename(file_temp1, file1)
    if(file_temp2 != file2):
        os.rename(file_temp2, file2)


def compara():
    # FUNÇÃO RESPONSAVEL PELA COMPARAÇÃO DAS DUAS TABELAS
    global table_novas
    global table_excluidas
    global table_discrep
    global table_discrep1
    global table1
    global table2
    global selected_table
    global campos
    # Copia os nomes das colunas das tabelas carregadas
    # para os 3 dataframes do relatório
    # table_novas = table1[0:0]
    # table_excluidas = table1[0:0]
    # table_discrep = table1[0:0]

    # Preenche o dataframe table_excluidas com as
    # linhas que possuem valores de 'RTUNO' e 'PNTNO'
    # que existem na tabela1 e não na tabela2
    table_excluidas = table1[~table1.set_index(campos).index
                             .isin(table2.set_index(campos).index)]

    # Preenche o dataframe table_excluidas com as
    # linhas que possuem valores de 'RTUNO' e 'PNTNO'
    # que existem na tabela2 e não na tabela1
    table_novas = table2[~table2.set_index(campos).index.
                         isin(table1.set_index(campos).index)]

    # Preenche o dataframe table_discrep1 com as
    # linhas da tabela 1que possuem a combinação
    # de 'RTUNO' e 'PNTNO' e estão presentes nos dois bancos
    table_discrep1 = table1[table1.set_index(campos).index.
                            isin(table2.set_index(campos).index)]
    # Mantem no dataframe auxiliar as linhas
    # que possuem todos as colunas iguais na tabela2
    table_aux = table1[table1.set_index(table1.columns.tolist()).
                       index.isin(table2.set_index(
                           table1.columns.tolist()).index)]

    # Mantem no dataframe table_discrep1 somente as
    # linhas que estão no datafame table_discrep1 e
    # não estão no dataframe table_aux, gerando assim
    # somente as linhas que possuem valores diferentes
    # Não é o melhor jeito de fazer mas é o que funciona
    # Tentar tirar as linhas que são iguais esvazia o
    # dataframe na primeira iteração
    table_discrep1 = table_discrep1[~table_discrep1.set_index(
        campos).index.isin(table_aux.set_index(campos).index)]

    # Preenche o dataframe table_discrep2 com as linhas
    # da tabela 2 que possuem a combinação
    # de 'RTUNO' e 'PNTNO' e estão presentes nos dois bancos
    table_discrep2 = table2[table2.set_index(campos).index.
                            isin(table1.set_index(campos).index)]

    table_aux = table2[table2.set_index(
        table1.columns.tolist()).index.isin(
            table1.set_index(table1.columns.tolist()).index)]

    # Mantem no dataframe table_discrep1 somente as linhas
    # que estão no datafame table_discrep1 e
    # não estão no dataframe table_aux, gerando assim somente
    # as linhas que possuem valores diferentes
    # Não é o melhor jeito de fazer mas é o que funciona
    # Tentar tirar as linhas que são iguais esvazia o
    # dataframe na primeira iteração
    table_discrep2 = table_discrep2[~table_discrep2.set_index(
        campos).index.isin(table_aux.set_index(campos).index)]

    # Verifica se os dataframes que indicam as linhas
    # discrepantes em cada tabela possuem o mesmo tamanho
    if(table_discrep1.shape[0] == table_discrep2.shape[0]):
        # Insere uma coluna no inicio dos dataframes de
        # discrepancias com o nome do arquivo original de cada linha
        path_1 = path1
        path_2 = path2
        while(path_1.find("/") != -1):
            path_1 = path_1[1:]
        while(path_2.find("/") != -1):
            path_2 = path_2[1:]
        table_discrep1.insert(loc=0, column='Arquivo', value=path_1)
        table_discrep2.insert(loc=0, column='Arquivo', value=path_2)
        # Copia os indices do dataframe table_discrep2
        # para adicionar a coluna Arquivo
        table_discrep = table_discrep2[0:0]
        # Itera pelos dataframes de discrepancia para
        # organizar o dataframe table_discrep
        # Adiciona uma linha da table_discrep1 por vez
        # e procura na table_discrep2
        # a linha com os mesmos valores de 'RTUNO','PNTNO'
        # para adicionar na sequencia
        # AFETA MUITO O DESEMPENHO POREM FACILITA A VISUALIZAÇÃO
        for i in range(table_discrep1.shape[0]):
            table_discrep = pd.concat([table_discrep, table_discrep1.
                                       iloc[[i]]], ignore_index=False)
            for j in range(table_discrep2.shape[0]):
                if(table_discrep2.iat[j, 1] == table_discrep1.iat[i, 1]
                   and table_discrep2.iat[j, 2] == table_discrep1.iat[i, 2]):
                    table_discrep = pd.concat([table_discrep, table_discrep2.
                                               iloc[[j]]], ignore_index=False)
                    break
    else:
        # Se não tiverem o mesmo tamanho alguma coisa errada aconteceu
        print("ACONTECEU ALGUMA COISA ERRADA NA PARTE DAS LINHAS DISCREPANTES")

    # Devido ao fato do dataframe iniciar com indice 0
    # e a tabela do pandastable com indice 1
    # É necessário aplicar uma correção nos indices dos 3 dataframes:

    # Correção table_discrep
    temp_index = table_discrep.index.tolist()
    for i in range(len(temp_index)):
        temp_index[i] = temp_index[i]+1
    table_discrep.index = temp_index

    # Correção table_novas
    temp_index = table_novas.index.tolist()
    for i in range(len(temp_index)):
        temp_index[i] = temp_index[i]+1
    table_novas.index = temp_index

    # Correção table_excluidas
    temp_index = table_excluidas.index.tolist()
    for i in range(len(temp_index)):
        temp_index[i] = temp_index[i]+1
    table_excluidas.index = temp_index

    # Função que atualiza as tabelas na interface
    update_table(0)


###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################
###########################################################

# CRIA A JANELA PRINCIPAL
root = tk.Tk()
# Variáveis com a resolução da tela para ajustar a posição das tabelas
width = root.winfo_screenwidth()
height = root.winfo_screenheight()
# Faz com que a janela principal tenha o tamanho igual a resolução
root.geometry("%dx%d" % (width, height))
root.title("COMPARADOR ACCESS v1.5.2")
# Maximiza a janela principal
root.state("zoomed")

if __name__ == '__main__':
    multiprocessing.freeze_support()

    def myinfo():
        # Função que mostra as informações do algoritmo
        str_info = "Autor: Rafael Henrique da Rosa\n"
        str_info += "Supervisor: Mauricio Menon\n"
        str_info += "Estagiário Itaipu Binacional- SMIN.DT - Abril de 2022\n"
        str_info += "O algoritmo compara duas tabelas em arquivos access "
        str_info += "e exibe linhas "
        str_info += "excluidas, novas e discrepantes."

        messagebox.showinfo("Info", str_info)

    def show_tutorial():
        # Função que exibe um pequeno tutorial
        # Abre o arquivo 'tutorial.txt' que deve estar na pasta do algoritmo
        f = open(resource_path("tutorial.txt"), "rt", encoding='utf-8')
        x = f.read()
        # Mostra o conteudo do arquivo em uma messagebox
        messagebox.showinfo("Info", x)

    def close_root():
        # Função para confimar o fechamento da interface
        if messagebox.askyesno("SAIR", "Fechar o aplicativo?"):
            root.destroy()

    # dataframe temporário para exibir linhas em branco ao iniciar o programa
    # Unicamente estético, não altera performance
    df = pd.DataFrame({
        'A': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'B': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'C': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'D': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'E': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'F': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'G': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'H': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'I': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'J': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'K': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'L': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'M': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'N': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'O': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'P': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'Q': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'R': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'S': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'T': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'U': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
    })

    def unfilter():
        btn.place_forget()
        label3.place_forget()
        update_table(0)

    btn = ttk.Button(root, text="X", command=unfilter)
    btn.place(x=width-200-22, y=50, height=20, width=20)
    label3 = ttk.Label(text="Filtro: RTUNO")
    label3.place(x=width-200, y=50, height=20, width=200)
    btn.place_forget()
    label3.place_forget()
    # PARTE DA PESQUISA

    def filtra():
        global campo_pesquisa
        global table1_filtrada
        global table2_filtrada
        global table_discrep_filtrada
        global table_excluidas_filtrada
        global table_novas_filtrada

        table1_filtrada = table1.astype(str)
        table2_filtrada = table2.astype(str)
        table_discrep_filtrada = table_discrep.astype(str)
        table_excluidas_filtrada = table_excluidas.astype(str)
        table_novas_filtrada = table_novas.astype(str)

        table1_filtrada = table1_filtrada[
            table1_filtrada[campo_pesquisa].str.contains(
                texto_pesquisa, flags=re.IGNORECASE)]
        table2_filtrada = table2_filtrada[table2_filtrada[
            campo_pesquisa].str.contains(texto_pesquisa, flags=re.IGNORECASE)]
        table_discrep_filtrada = table_discrep_filtrada[
            table_discrep_filtrada[campo_pesquisa].str.contains(
                texto_pesquisa, flags=re.IGNORECASE)]
        table_excluidas_filtrada = table_excluidas_filtrada[
            table_excluidas_filtrada[campo_pesquisa].str.contains(
                texto_pesquisa, flags=re.IGNORECASE)]
        table_novas_filtrada = table_novas_filtrada[
            table_novas_filtrada[campo_pesquisa].str.contains(
                texto_pesquisa, flags=re.IGNORECASE)]
        update_table(1)
        label3.config(text="FILTRO: " + texto_pesquisa)
        btn.place(x=width-200-22, y=50, height=20, width=20)
        label3.place(x=width-200, y=50, height=20, width=200)

    def find():

        if table1.empty or table2.empty:
            messagebox.showinfo(
                "ERRO", "Banco de dados não selecionados")
        else:
            colunas_pesquisar = colunas
            colunas_pesquisar.pop(0)
            # colunas_pesquisar.insert(0, "TODOS")
            child_w = Toplevel(root)
            child_w.geometry("450x110")
            child_w.grab_set()
            child_w.title("FILTRAGEM")
            Frm = Frame(child_w)
            Label(Frm, text='Enter Word to Find:')
            # Label.place(x=0, y=50, height=250, width=750)
            Frm.place(x=0, y=0, height=110, width=450)
            modu = Entry(Frm)
            modu.place(x=10, y=30, height=30, width=200)
            modu.focus_set()

            buttn = ttk.Button(Frm, text='FILTRAR')
            buttn.place(x=450/2-40, y=70, height=35, width=80)
            # Create Label in Mainwindow and Childwindow
            label_child = Label(child_w, text="Pesquisar por:")
            label_child.place(x=10+200/2-40, y=0, height=20, width=80)

            label_child2 = Label(child_w, text="Nos campos:")
            label_child2.place(x=450/2+80, y=0, height=20, width=80)

            selected = tk.StringVar()
            c2_cb = ttk.Combobox(child_w, width=50, textvariable=selected)
            c2_cb['values'] = colunas_pesquisar
            c2_cb['state'] = 'readonly'
            c2_cb.pack(fill=tk.X, padx=5, pady=5)
            c2_cb.place(x=450/2+15, y=30, height=30, width=200)
            # c2_cb.current(0)

            def xx():
                global texto_pesquisa
                global campo_pesquisa
                texto_pesquisa = modu.get()
                campo_pesquisa = selected.get()
                child_w.destroy()
                filtra()

            buttn.config(command=xx)

    def select_campos():
        global campos

        def reset_campos():
            while len(campos) != 3:
                campos.append('Nenhum')

        # # Quando a tabela é selecionada executa a comparação
        # def select_1(event):
        #     global campos
        #     campos[0] = selected_1.get()

        # def select_2(event):
        #     global campos
        #     campos[1] = selected_2.get()

        # def select_3(event):
        #     global campos
        #     campos[2] = selected_3.get()

        def try_compara():
            global campos
            reset_campos()
            # campos = ['Nenhum', 'Nenhum', 'Nenhum']

            campos[0] = selected_1.get()
            campos[1] = selected_2.get()
            campos[2] = selected_3.get()

            if(campos[0] != 'Nenhum' or campos[1] != 'Nenhum'
               or campos[2] != 'Nenhum'):
                campos = list(dict.fromkeys(campos))

                for i in range(len(campos)):
                    if campos[i] == 'Nenhum':
                        del campos[i]

                compara()
            else:
                messagebox.showinfo(
                    "ERRO", "SELECIONE UM CAMPO PARA COMPARAÇÃO")

        # Cria uma label para indicar que a tabela deve ser selecionada
        label = ttk.Label(text="Selecione os campos de comparação:")
        label.place(x=400, y=0, height=20, width=250)
        global colunas
        colunas = table1.columns.tolist()
        colunas.insert(0, 'Nenhum')

        selected_1 = tk.StringVar()
        c1_cb = ttk.Combobox(root, width=50, textvariable=selected_1)
        c1_cb['values'] = colunas
        c1_cb['state'] = 'readonly'
        c1_cb.pack(fill=tk.X, padx=5, pady=5)
        c1_cb.place(x=400, y=30, height=30, width=200)
        c1_cb.current(0)

        selected_2 = tk.StringVar()
        c2_cb = ttk.Combobox(root, width=50, textvariable=selected_2)
        c2_cb['values'] = colunas
        c2_cb['state'] = 'readonly'
        c2_cb.pack(fill=tk.X, padx=5, pady=5)
        c2_cb.place(x=630, y=30, height=30, width=200)
        c2_cb.current(0)

        selected_3 = tk.StringVar()
        c3_cb = ttk.Combobox(root, width=50, textvariable=selected_3)
        c3_cb['values'] = colunas
        c3_cb['state'] = 'readonly'
        c3_cb.pack(fill=tk.X, padx=5, pady=5)
        c3_cb.place(x=630+230, y=30, height=30, width=200)
        c3_cb.current(0)

        botao_compara = ttk.Button(root, text='COMPARAR', command=try_compara)
        botao_compara.place(x=630+230+230, y=30, height=30, width=130)

        # c1_cb.bind('<<ComboboxSelected>>', select_1)
        # c2_cb.bind('<<ComboboxSelected>>', select_2)
        # c3_cb.bind('<<ComboboxSelected>>', select_3)

    def select_table(file_type):
        # Função para selecionar a tabela
        global table_obj
        # Variável para uma lista das tabelas presentes no arquivo antigo
        global output_tables
        global selected_table
        selected_table = ""

        # Caso seja arquivo access:
        if file_type == 'access':
            # Cria a linha de comando no cmd que executa o arquivo mdb-tables
            # e guarda o output na lista
            path = resource_path('mdbtools\\mdb-tables.exe')
            output_tables = subprocess.check_output(
                [path, path1]).decode()
            output_tables = output_tables.split()

            path = resource_path('mdbtools\\mdb-tables.exe')
            output_tables2 = subprocess.check_output(
                [path, path2]).decode()
            output_tables2 = output_tables2.split()

        # Caso seja arquivo excel:
        elif file_type == 'excel':
            try:
                table_obj = openpyxl.load_workbook(path1)
                table_obj2 = openpyxl.load_workbook(path2)

            except openpyxl.utils.exceptions.InvalidFileException:
                print("ai realmente não ta abrindo o arquivo antigo")
            # Guarda o nome dos sheets na lista
            output_tables = table_obj.sheetnames
            output_tables2 = table_obj2.sheetnames

        # Cria uma label para indicar que a tabela deve ser selecionada
        label = ttk.Label(text="Selecione a tabela para comparar:")
        label.place(x=50, y=0, height=20, width=200)

        # Cria um combobox com a lista de tabelas
        selected_month = tk.StringVar()
        month_cb = ttk.Combobox(root, width=50, textvariable=selected_month)
        month_cb['values'] = output_tables
        month_cb['state'] = 'readonly'
        month_cb.pack(fill=tk.X, padx=5, pady=5)
        month_cb.place(x=50, y=30, height=30, width=200)

        # Quando a tabela é selecionada executa a comparação
        def month_changed(event):
            global selected_table
            selected_table = selected_month.get()
            if selected_table in output_tables2:
                load_tables()
                select_campos()
            else:
                messagebox.showinfo(
                    "ERRO", "TABELA NÃO EXISTE NO ARQUIVO NOVO")
        month_cb.bind('<<ComboboxSelected>>', month_changed)

    def select_file_access2():
        # Função para seleção do banco novo access
        global path1
        global path2

        while(True):
            file_types = (('Access Files', '*.accdb'), ('All files', '*.*'))
            file_name = fd.askopenfilename(
                title='SELECIONAR ARQUIVO NOVO', filetypes=file_types)

            path2 = file_name
            if file_name.endswith('.accdb') is False:
                messagebox.showinfo
                ("ERRO", "SELECIONE UM ARQUIVO ACCESS (.accdb)")
                break
            if path1 == path2:
                messagebox.showinfo("ERRO", "ARQUIVOS SELECIONADOS IGUAIS")
            else:
                break

        if(path1 != "" and path2 != ""):
            # Chama a função para selecionar a tabela
            select_table('access')

    def select_file_access():
        # Função de seleção do banco antigo access
        global path1
        global path2
        clear_table()
        while(True):
            file_types = (('Access Files', '*.accdb'), ('All files', '*.*'))
            file_name = fd.askopenfilename(title='SELECIONAR ARQUIVO ANTIGO',
                                           filetypes=file_types)

            path1 = file_name
            if file_name.endswith('.accdb') is False:
                messagebox.showinfo(
                    "ERRO", "SELECIONE UM ARQUIVO ACCESS (.accdb)")
                break
            else:
                break
            # Chama a função para selecionar o banco novo
        if path1 != "":
            select_file_access2()

    def select_file_excel2():
        # Função para seleção do banco novo excell
        global path1
        global path2

        while(True):
            file_types = (('Excel Files', '*.xlsx'), ('All files', '*.*'))
            file_name = fd.askopenfilename(
                title='SELECIONAR ARQUIVO NOVO', filetypes=file_types)

            path2 = file_name
            if file_name.endswith('.xlsx') is False:
                messagebox.showinfo(
                    "ERRO", "SELECIONE UM ARQUIVO EXCEL (.xlsx)")
                break
            if path1 == path2:
                messagebox.showinfo("ERRO", "ARQUIVOS SELECIONADOS IGUAIS")
            else:
                break

        if(path1 != "" and path2 != ""):
            # Chama a função para selecionar a tabela
            select_table('excel')

    def select_file_excel():
        # Função de seleção do banco antigo excell
        global path1
        global path2
        clear_table()
        while(True):
            file_types = (('Excel Files', '*.xlsx'), ('All files', '*.*'))
            file_name = fd.askopenfilename(title='SELECIONAR ARQUIVO ANTIGO',
                                           filetypes=file_types)

            path1 = file_name
            if file_name.endswith('.xlsx') is False:
                messagebox.showinfo(
                    "ERRO", "SELECIONE UM ARQUIVO EXCEL (.xlsx)")
                break
            else:
                break
            # Chama a função para selecionar o banco novo
        if path1 != "":
            select_file_excel2()

    def organiza_relat(file_path, is_complet):
        # Função que organiza os arquivos excel exportados

        global table_obj
        global table_discrep

        # definições de cores e borda
        cinza = PatternFill(start_color='787878', end_color='787878',
                            fill_type='solid')
        vermelho = PatternFill(start_color='ff0000', end_color='ff0000',
                               fill_type='solid')
        vermelho_claro = PatternFill(start_color='fa9696', end_color='fa9696',
                                     fill_type='solid')
        verde = PatternFill(start_color='82e89d', end_color='82e89d',
                            fill_type='solid')
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Abre o arquivo exportado
        try:
            table_obj = openpyxl.load_workbook(file_path)

        except openpyxl.utils.exceptions.InvalidFileException:
            try:
                file_path = file_path.replace("/", "\\")
                table_obj = openpyxl.load_workbook(file_path)

            except openpyxl.utils.exceptions.InvalidFileException:
                print("ai realmente não ta abrindo o arquivo antigo")

        # carrega o sheet
        table_sheet_resul_obj = table_obj['RELATÓRIO']
        if is_complet:
            table_sheet_antigo_obj = table_obj[table_obj.sheetnames[0]]
            table_sheet_novo_obj = table_obj[table_obj.sheetnames[1]]

            for i in range(1, table1.shape[1]):
                table_sheet_antigo_obj.cell(1, i).fill = cinza
                table_sheet_novo_obj.cell(1, i).fill = cinza

        # Adiciona labels no sheet de relatório
        table_sheet_resul_obj.cell(3, 1).value = "LINHAS DISCREPANTES:"
        table_sheet_resul_obj.cell(3, 1).font = Font(bold=True)
        table_sheet_resul_obj.cell(
            3+table_discrep.shape[0]+1+3, 1).value = "LINHAS ADICIONADAS "
        "(presentes somente no arquivo novo):"
        table_sheet_resul_obj.cell(
            3+table_discrep.shape[0] + 1 + 3, 1).font = Font(bold=True)
        table_sheet_resul_obj.cell(
            3+table_discrep.shape[0] + 1 + 3 +
            table_novas.shape[0]+3+1, 1).value = ""
        "LINHAS EXCLUIDAS(presentes somente no arquivo antigo):"
        table_sheet_resul_obj.cell(
            3+table_discrep.shape[0]+1+3 +
            table_novas.shape[0]+3+1, 1).font = Font(bold=True)
        # Retira o grid do relatório
        table_sheet_resul_obj.sheet_view.showGridLines = False

        # Pinta o header do relatorio de cinza
        for i in range(2, table_discrep.shape[1]+2):
            table_sheet_resul_obj.cell(4, i).fill = cinza
        for i in range(2, table_novas.shape[1]+2):
            table_sheet_resul_obj.cell(
                4+table_discrep.shape[0]+3+1, i).fill = cinza
            table_sheet_resul_obj.cell(
                4+table_discrep.shape[0] +
                3+table_novas.shape[0]+5, i).fill = cinza

        # Pinta os indices do relatório de cinza
        for i in range(0, table_discrep.shape[0]):
            table_sheet_resul_obj.cell(i+5, 1).fill = cinza
        for i in range(0, table_novas.shape[0]):
            table_sheet_resul_obj.cell(
                i+table_discrep.shape[0]+9, 1).fill = cinza
        for i in range(0, table_excluidas.shape[0]):
            table_sheet_resul_obj.cell(
                i+table_discrep.shape[0] + 8
                + table_novas.shape[0]+5, 1).fill = cinza

        # Corrige os idices do relatório devido ao dataframe iniciar em 0
        for i in range(5, table_discrep.shape[0]+5):
            table_sheet_resul_obj.cell(
                i, 1).value = table_sheet_resul_obj.cell(
                i, 1).value + 1
        for i in range(table_discrep.shape[0]+5+4,
                       table_discrep.shape[0]+5+4+table_novas.shape[0]):
            table_sheet_resul_obj.cell(
                i, 1).value = table_sheet_resul_obj.cell(i, 1).value + 1
        for i in range(table_discrep.shape[0]+5+4+table_novas.shape[0]+4,
                       table_discrep.shape[0]+5+4+table_novas.shape[0] +
                       4+table_excluidas.shape[0]):
            table_sheet_resul_obj.cell(
                i, 1).value = table_sheet_resul_obj.cell(i, 1).value + 1

        # se for uma exportação completa com os 3 sheets

        # Pinta as linhas discrepantes de vermelho
        for i in range(2, table_discrep.shape[1]+2):
            for j in range(5, table_discrep.shape[0]+5):
                table_sheet_resul_obj.cell(j, i).border = borda_fina

                if j % 2 == 1:
                    if i != 2:
                        if table_sheet_resul_obj.cell(
                                j, i).value != table_sheet_resul_obj.cell(
                                    j+1, i).value:
                            table_sheet_resul_obj.cell(
                                j, i).fill = vermelho
                            table_sheet_resul_obj.cell(
                                j+1, i).fill = vermelho
                            if is_complet:
                                for k in range(1, table_discrep.shape[1]):
                                    table_sheet_antigo_obj.cell(
                                        table_sheet_resul_obj.cell(
                                            j, 1).value, k).fill = vermelho
                                    table_sheet_novo_obj.cell(
                                        table_sheet_resul_obj.cell(
                                            j+1, 1).value,
                                        k).fill = vermelho

        # Pinta as linhas novas de vermelho claro
        for i in range(2, table_novas.shape[1]+2):
            for j in range(table_discrep.shape[0]+5+4,
                           table_discrep.shape[0] + 5
                           + 4 + table_novas.shape[0]):
                table_sheet_resul_obj.cell(j, i).border = borda_fina
                table_sheet_resul_obj.cell(j, i).fill = vermelho_claro
                if is_complet:
                    for k in range(1, table_novas.shape[1]):
                        table_sheet_novo_obj.cell(
                            table_sheet_resul_obj.cell(
                                j, 1).value, k).fill = vermelho_claro

            # Pinta as linhas excluidas de verde
            for i in range(2, table_excluidas.shape[1]+2):
                for j in range(
                        table_discrep.shape[0] + 9 + table_novas.shape[0] +
                        4, table_discrep.shape[0] + 5 + 4 +
                        table_novas.shape[0] + 4 + table_excluidas.shape[0]):
                    table_sheet_resul_obj.cell(j, i).border = borda_fina
                    table_sheet_resul_obj.cell(j, i).fill = verde
                    if is_complet:
                        for k in range(1, table_excluidas.shape[1]):
                            table_sheet_antigo_obj.cell(
                                table_sheet_resul_obj.cell(
                                    j, 1).value, k).fill = verde

        table_obj.save(file_path)

    def select_file_export_Antiga():
        # Função que exporta a tabela antiga para um arquivo Excel xlsx
        global selected_table
        if selected_table == "":
            messagebox.showinfo("Info", "Nenhuma tabela selecionada")
        else:
            file_types = (('Excel files', '*.xlsx'), ('All files', '*.*'))
            file_path = tk.filedialog.asksaveasfilename(
                title='SALVAR TABELA ANTIGA', filetypes=file_types)
            if file_path.endswith('.xlsx') is False:
                file_path += '.xlsx'
            table1.to_excel(file_path,
                            sheet_name=selected_table, index=False)
            str_temp = "start EXCEL.EXE " + file_path
            os.system(str_temp)

    def select_file_export_Nova():
        # Função que exporta a tabela nova para um arquivo Excel xlsx
        global selected_table
        global table1
        if selected_table == "":
            messagebox.showinfo("Info", "Nenhuma tabela selecionada")
        else:
            file_types = (('Excel files', '*.xlsx'), ('All files', '*.*'))
            file_path = tk.filedialog.asksaveasfilename(
                title='SALVAR TABELA NOVA', filetypes=file_types)
            if file_path.endswith('.xlsx') is False:
                file_path += '.xlsx'
            table2.to_excel(file_path,
                            sheet_name=selected_table, index=False)
            str_temp = "start EXCEL.EXE " + file_path
            os.system(str_temp)

    def select_file_export_Relat():
        # Função que exporto relatório para um arquivo Excel xlsx
        global selected_table
        if selected_table == "":
            messagebox.showinfo("Info", "Nenhuma tabela selecionada")
        else:
            file_types = (('Excel files', '*.xlsx'), ('All files', '*.*'))
            file_path = tk.filedialog.asksaveasfilename(
                title='SALVAR TABELA NOVA',
                filetypes=file_types)
            if file_path.endswith('.xlsx') is False:
                file_path += '.xlsx'
            writer = pd.ExcelWriter(file_path, engine='xlsxwriter')

            def multiple_dfs(df_list, sheets, file_name, spaces):
                row = 3
                for dataframe in df_list:
                    dataframe.to_excel(writer, sheet_name=sheets,
                                       startrow=row, startcol=0,
                                       index=True)
                    row = row + len(dataframe.index) + spaces + 1
                writer.save()

            # list of dataframes
            dfs = [table_discrep, table_novas, table_excluidas]

            # run function
            multiple_dfs(dfs, 'RELATÓRIO', file_path, 3)
            try:
                writer.close()
            except Exception:
                pass
            organiza_relat(file_path, 0)
            str_temp = "start EXCEL.EXE " + file_path
            os.system(str_temp)

    def select_file_export_Complet():
        # Função que exporta tudo para um arquivo Excel xlsx
        global selected_table
        global table1
        global path1, path2
        path_1 = path1
        path_2 = path2
        if selected_table == "":
            messagebox.showinfo("Info", "Nenhuma tabela selecionada")
        else:
            file_types = (('Excel files', '*.xlsx'), ('All files', '*.*'))
            file_path = tk.filedialog.asksaveasfilename(
                title='SALVAR TABELA NOVA', filetypes=file_types)
            if file_path.endswith('.xlsx') is False:
                file_path += '.xlsx'
            # Tira o caminho e deixa só o nome dos arquivos selecionados
            while(path_1.find("/") != -1):
                path_1 = path_1[1:]
            while(path_2.find("/") != -1):
                path_2 = path_2[1:]
            if len(path_1) >= 31 or len(path_2) >= 31:
                path_1 = "ANTIGO"
                path_2 = "NOVO"
            writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
            table1.to_excel(writer,
                            sheet_name=path_1, index=False)
            table2.to_excel(writer,
                            sheet_name=path_2, index=False)

            def multiple_dfs(df_list, sheets, file_name, spaces):
                row = 3
                for dataframe in df_list:
                    dataframe.to_excel(writer, sheet_name=sheets,
                                       startrow=row, startcol=0, index=True)
                    row = row + len(dataframe.index) + spaces + 1
                writer.save()

            # list of dataframes
            dfs = [table_discrep, table_novas, table_excluidas]

            # run function
            multiple_dfs(dfs, 'RELATÓRIO', file_path, 3)
            try:
                writer.close()
            except Exception:
                pass
            organiza_relat(file_path, 1)
            str_temp = "start EXCEL.EXE " + file_path
            os.system(str_temp)

    # Adiciona um menu a janela principal
    menubar = tk.Menu(root)

    filemenu = tk.Menu(menubar, tearoff=0)
    filemenu.add_command(label="Selecionar Arquivo Access (.accdb)",
                         command=select_file_access)
    filemenu.add_command(label="Selecionar Arquivo Excel (.xlsx)",
                         command=select_file_excel)
    filemenu.add_command(label="Sair", command=close_root)
    helpmenu = tk.Menu(menubar, tearoff=0)
    helpmenu.add_command(label="Como usar", command=show_tutorial)
    helpmenu.add_command(label="Sobre o programa", command=myinfo)
    exportmenu = tk.Menu(menubar, tearoff=0)
    exportmenu.add_command(label="Exportar CSV tabela antiga",
                           command=select_file_export_Antiga)
    exportmenu.add_command(label="Exportar CSV tabela nova",
                           command=select_file_export_Nova)
    exportmenu.add_command(label="Exportar CSV relatório",
                           command=select_file_export_Relat)
    exportmenu.add_command(label="Exportar CSV Completo",
                           command=select_file_export_Complet)
    optionsmenu = tk.Menu(menubar, tearoff=0)
    colore = tk.BooleanVar()
    colore.set(False)
    optionsmenu.add_checkbutton(label='Colorir Ocorrencias (BETA)',
                                onvalue=1, offvalue=0, variable=colore)
    optionsmenu.add_command(label="Filtrar", command=find)

    menubar.add_cascade(label="Arquivo", menu=filemenu)
    menubar.add_cascade(label="Exportar", menu=exportmenu)
    menubar.add_cascade(label="Opções", menu=optionsmenu)
    menubar.add_cascade(label="Ajuda", menu=helpmenu)
    root.config(menu=menubar)

    # Cria 3 abas na janela principal para exibir as tabelas e o relatório
    tabControl = ttk.Notebook(root)
    tabControl.place(x=0, y=70, height=height, width=width)
    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)
    tab3 = ttk.Frame(tabControl)
    tabControl.add(tab1, text='RELATÓRIO')
    tabControl.add(tab2, text='ARQUIVO ANTIGO')
    tabControl.add(tab3, text='ARQUIVO NOVO')

    #  from tkinter import BOTH, LEFT,RIGHT
    # container = Frame(tabControl)
    # container.place(x=0, y=100, height=height, width=width)
    # canvas = tk.Canvas(container, width=width, height=height)
    # canvas.place(x=0, y=70, height=height, width=width)
    # scroll = tk.Scrollbar(container, command=canvas.yview)
    # canvas.config(yscrollcommand=scroll.set, scrollregion=(0,0,100,1000))
    # canvas.pack(side=LEFT, fill=BOTH, expand=True)
    # scroll.pack(side=RIGHT, fill=tk.Y)

    canvas1 = tk.Canvas(tab1, width=width, height=height)
    # scroll = tk.Scrollbar(tab1, command=canvas1.yview)
    # canvas1.config(yscrollcommand=scroll.set, scrollregion=(0,0,0,1500))
    canvas1.place(x=0, y=0, height=height, width=width)
    # scroll.place(x=width-20, y=0, height=height, width=16)

    # Adiciona o frame da tabela antiga na aba 'arquivo antigo'
    frameOne = Frame(canvas1, width=width, height=450)
    canvas1.create_window(0, 0, anchor=tk.NW,
                          window=frameOne, width=width, height=height)

    frame1 = tk.Frame(tab2)
    frame1.place(x=0, y=0, height=height-178, width=width)
    pt1 = Table(frame1)
    pt1.model.df = df
    pt1.autoResizeColumns()
    pt1.show()
    pt1.autoResizeColumns()
    pt1.redraw()

    # Adiciona o frame da tabela nova na aba 'arquivo novo'
    frame2 = tk.Frame(tab3)
    frame2.place(x=0, y=0, height=height-178, width=width)
    pt2 = Table(frame2)
    pt2.model.df = df
    pt2.autoResizeColumns()
    pt2.show()
    pt2.autoResizeColumns()
    pt2.autoResizeColumns()
    pt2.redraw()
    factor = 3.3
    # Label das linhas discrepantes
    lbl_discrep = ttk.Label(frameOne, text="LINHAS DISCREPANTES:",
                            font='Helvetica 12 bold')
    lbl_discrep.place(x=0, y=0, height=22, width=width-30)

    # Adiciona um frame para exibir as linhas discrepantes
    frame_resul_discrep = tk.Frame(frameOne)
    frame_resul_discrep.place(x=0, y=20,
                              height=(height/factor)-(22*3), width=width)
    pt_resul_discrep = Table(frame_resul_discrep)
    pt_resul_discrep.model.df = df
    options = {
        'cellbackgr': '#f7f6dc',
        # 'rowselectedcolor': '#f7f6dc',
        'textcolor': 'black'}
    config.apply_options(options, pt_resul_discrep)
    pt_resul_discrep.show()
    pt_resul_discrep.autoResizeColumns()
    pt_resul_discrep.show()
    pt_resul_discrep.redraw()

    # Label das linhas novas
    lbl_novas = ttk.Label(frameOne, text="LINHAS ADICIONADAS "
                          "(presentes somente no arquivo novo):",
                          font='Helvetica 12 bold')
    lbl_novas.place(x=0, y=(height/factor)+25-(22*3),
                    height=22, width=width)

    # Adiciona um frame para exibir as linhas novas
    frame_resul_novas = tk.Frame(frameOne)
    frame_resul_novas.place(x=0, y=(height/factor)+25+25-(22*3),
                            height=(height/factor)-(22*3), width=width)
    pt_resul_novas = Table(frame_resul_novas)
    pt_resul_novas.model.df = df
    options = {
        'rowselectedcolor': '#cf4040',
        'cellbackgr': '#cf4040',
        'textcolor': 'black'}
    config.apply_options(options, pt_resul_novas)
    pt_resul_novas.show()
    pt_resul_novas.autoResizeColumns()
    pt_resul_novas.show()
    pt_resul_novas.redraw()

    # Label das linhas excluidas
    lbl_excluidas = ttk.Label(frameOne,
                              text="LINHAS EXCLUIDAS "
                              "(presentes somente no arquivo antigo):",
                              font='Helvetica 12 bold')
    lbl_excluidas.place(x=0, y=((height/factor)*2+25+30)-(22*3*2),
                        height=22, width=width)

    # Adiciona um frame para exibir as linhas excluidas
    frame_resul_excluidas = tk.Frame(frameOne)
    frame_resul_excluidas.place(x=0, y=(height/factor)*2+25+25+30-(22*3*2),
                                height=(height/factor)-(22*3), width=width)

    pt_resul_excluidas = Table(frame_resul_excluidas)

    pt_resul_excluidas.model.df = df
    options = {
        'cellbackgr': '#baf5c3',
        'rowselectedcolor': '#baf5c3',
        # 'colheadercolor': '#f71616',
        'gridcolor': 'black',
        'textcolor': 'black'}
    config.apply_options(options, pt_resul_excluidas)
    pt_resul_excluidas.autoResizeColumns()
    pt_resul_excluidas.show()
    pt_resul_excluidas.redraw()

    root.protocol("WM_DELETE_WINDOW", close_root)

    image_path = resource_path("icone.ico")
    root.iconbitmap(image_path)

    def resize(event):
        global width, height
        global scroll, canvas1
        if (width != root.winfo_width()) and (height != root.winfo_width()):
            width = event.width
            height = event.height
            factor = 2.95
            lbl_discrep.place(x=0, y=0, height=22, width=width)
            frame_resul_discrep.place(x=0, y=20,
                                      height=(height/factor)-(22*3),
                                      width=width)

            lbl_novas.place(x=0, y=(height/factor)+25-(22*3),
                            height=22, width=width)
            frame_resul_novas.place(x=0, y=(height/factor)+25+25-(22*3),
                                    height=(height/factor)-(22*3), width=width)

            lbl_excluidas.place(x=0, y=((height/factor)*2+25+30)-(22*3*2),
                                height=22, width=width)
            frame_resul_excluidas.place(x=0,
                                        y=(height/factor)*2+25+25+30-(22*3*2),
                                        height=(height/factor)-(22*3),
                                        width=width)

            frame1.place(x=0, y=0, height=height-100, width=width)
            frame2.place(x=0, y=0, height=height-100, width=width)
            # scroll.place(x=width-20, y=0, height=height, width=16)
            # print(
            # f"The width of Toplevel is {width} and the height of Toplevel "
            #      f"is {height}")

    root.bind("<Configure>", resize)
    # Loop janela principal
    root.mainloop()
